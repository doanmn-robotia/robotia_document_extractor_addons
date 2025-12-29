# -*- coding: utf-8 -*-
from odoo import http, _
from odoo.http import request
import base64
import json
import logging
import time

_logger = logging.getLogger(__name__)


class ExtractionController(http.Controller):
    """
    JSON-RPC Controller for document extraction

    IMPORTANT: This controller does NOT create records immediately.
    It extracts data and returns an action to open form in CREATE mode.
    """

    @staticmethod
    def _calculate_avg_quantity(usage_record):
        """
        Calculate correct average quantity from 3-year data

        FIX: Divide by count of non-null years, not always 3

        Args:
            usage_record: substance.usage record with year_1/2/3 fields

        Returns:
            tuple: (avg_kg, avg_co2)
        """
        # If avg fields already computed, use them
        if usage_record.avg_quantity_kg is not None and usage_record.avg_quantity_co2 is not None:
            return (usage_record.avg_quantity_kg, usage_record.avg_quantity_co2)

        # Count non-null years for kg
        kg_values = [
            usage_record.year_1_quantity_kg,
            usage_record.year_2_quantity_kg,
            usage_record.year_3_quantity_kg
        ]
        co2_values = [
            usage_record.year_1_quantity_co2,
            usage_record.year_2_quantity_co2,
            usage_record.year_3_quantity_co2
        ]

        kg_count = sum(1 for v in kg_values if v is not None)
        co2_count = sum(1 for v in co2_values if v is not None)

        # Calculate average (divide by actual count, not 3)
        avg_kg = sum(v for v in kg_values if v is not None) / kg_count if kg_count > 0 else 0
        avg_co2 = sum(v for v in co2_values if v is not None) / co2_count if co2_count > 0 else 0

        return (avg_kg, avg_co2)

    @http.route('/robotia/pdf_to_images', type='json', auth='user', methods=['POST'])
    def pdf_to_images(self, pdf_file):
        """
        Convert PDF to images and store as public attachments

        Args:
            pdf_file (str): Base64 encoded PDF

        Returns:
            dict: {
                'status': 'success',
                'pages': [
                    {
                        'attachment_id': 123,
                        'url': '/web/content/123',
                        'page_num': 0,
                        'filename': 'page_0.png'
                    },
                    ...
                ]
            }
        """
        try:
            # Decode PDF
            pdf_binary = base64.b64decode(pdf_file)

            # Convert to images
            ExtractionService = request.env['document.extraction.service'].sudo()
            image_paths = ExtractionService._pdf_to_images(pdf_binary)

            # Create attachments for each image
            Attachment = request.env['ir.attachment'].sudo()
            pages_metadata = []

            for idx, path in enumerate(image_paths):
                try:
                    # Read image file
                    with open(path, "rb") as image_file:
                        image_binary = image_file.read()

                    # Create public attachment
                    filename = f"page_{idx}.png"
                    attachment = Attachment.create({
                        'name': filename,
                        'type': 'binary',
                        'datas': base64.b64encode(image_binary).decode('utf-8'),
                        'res_model': 'document.extraction',
                        'res_id': 0,  # Temporary attachment
                        'public': True,  # Makes accessible via URL
                        'mimetype': 'image/png',
                        'description': f'PDF page preview {idx}',
                    })

                    # Build metadata
                    pages_metadata.append({
                        'attachment_id': attachment.id,
                        'url': f'/web/content/{attachment.id}',
                        'page_num': idx,
                        'filename': filename,
                    })

                    _logger.info(f"Created attachment {attachment.id} for page {idx}")

                finally:
                    # Clean up temp file
                    try:
                        import os
                        os.remove(path)
                    except Exception as cleanup_error:
                        _logger.warning(f"Failed to cleanup {path}: {cleanup_error}")

            return {
                'status': 'success',
                'pages': pages_metadata
            }

        except Exception as e:
            _logger.exception("Error converting PDF to images")
            return {'status': 'error', 'message': str(e)}

    @http.route('/robotia/extract_pages', type='json', auth='user', methods=['POST'])
    def extract_pages(self, attachment_ids, document_type='01', filename=None):
        """
        Create async extraction job for selected pages
        
        Flow:
        1. Merge selected page images into PDF
        2. Create ir.attachment for merged PDF
        3. Create extraction.job record
        4. Enqueue job with queue_job
        5. Cleanup temp page attachments
        6. Return success message
        
        Args:
            attachment_ids (list): List of attachment IDs (page images)
            document_type (str): '01' or '02'
            filename (str): Original filename
            
        Returns:
            dict: Success message for UI
        """
        try:
            # 1. Merge pages into single PDF
            service = request.env['document.extraction.service']
            merged_pdf_binary = service.merge_page_attachments_to_pdf(attachment_ids)
            
            # 2. Create attachment for merged PDF
            Attachment = request.env['ir.attachment'].sudo()
            merged_attachment = Attachment.create({
                'name': filename or f'merged_{document_type}.pdf',
                'type': 'binary',
                'datas': base64.b64encode(merged_pdf_binary).decode('utf-8'),
                'res_model': 'extraction.job',
                'mimetype': 'application/pdf',
                'description': f'Merged PDF for extraction job',
            })
            
            # 3. Create extraction job record
            ExtractionJob = request.env['extraction.job'].sudo()
            job_record = ExtractionJob.create({
                'attachment_id': merged_attachment.id,
                'document_type': document_type,
                'state': 'pending',
            })
            
            # 4. Update attachment to link to job
            merged_attachment.res_id = job_record.id
            
            # 5. Generate identity_key from attachment ID
            identity_key = f"extraction_{merged_attachment.id}"


            try:
                # 7. Cleanup temporary page attachments
                # 6. Enqueue job with queue_job
                job_record.with_delay(
                    channel='extraction',
                    max_retries=1,  # No retry - only run once
                    identity_key=identity_key,
                    description=f"Extract {filename or document_type}"
                ).run_extraction_async()
                
                _logger.info(f"Created extraction job {job_record.name} (ID: {job_record.id})")
            except Exception as e:
                _logger.warning(f"Could not cleanup temp attachments: {e}")
            
            # 8. Return success message (UI will redirect to dashboard)


            return {
                'type': 'success',
                'message': _('Extraction job created successfully'),
                'job_id': job_record.id,
                'job_name': job_record.name,
                'uuid': job_record.uuid
            }
            
        except Exception as e:
            _logger.exception("Failed to create extraction job")
            return {
                'type': 'error',
                'message': str(e),
            }
    
    @http.route('/robotia/get_my_extraction_jobs', type='json', auth='user', methods=['POST'])
    def get_my_extraction_jobs(self, offset=0, limit=10):
        """
        Get extraction jobs for current user with pagination
        
        Uses queue.job as the source of truth.
        Gets extraction.job data via queue.job.records field (similar to related_action).
        
        Args:
            offset (int): Starting position for pagination (default: 0)
            limit (int): Number of records to fetch (default: 10)
        
        Returns:
            dict: {
                'jobs': [...],      # List of job dicts
                'total': 123,       # Total count
                'has_more': True    # Whether there are more records
            }
        """
        try:
            QueueJob = request.env['queue.job'].sudo()
            
            # Get queue.job records for extraction jobs created by current user
            queue_domain = [
                ('user_id', '=', request.env.uid),
                ('model_name', '=', 'extraction.job'),
                ('method_name', '=', 'run_extraction_async')
            ]
            
            # Get total count
            total_count = QueueJob.search_count(queue_domain)
            
            # Get queue jobs with pagination
            queue_jobs = QueueJob.search(
                queue_domain,
                order='date_created desc',
                limit=limit,
                offset=offset
            )
            
            if not queue_jobs:
                return {
                    'jobs': [],
                    'total': total_count,
                    'has_more': False
                }
            
            # Build result list
            jobs_list = []
            for queue_job in queue_jobs:
                # Get extraction.job via records field (like related_action_open_record)
                extraction_job = queue_job.records
                
                # Skip if no extraction.job or wrong model
                if not extraction_job or extraction_job._name != 'extraction.job':
                    _logger.warning(f"Queue job {queue_job.uuid} has no valid extraction.job record")
                    continue
                
                # If multiple records, take the first one (shouldn't happen)
                if len(extraction_job) > 1:
                    _logger.warning(f"Queue job {queue_job.uuid} has multiple extraction.job records, using first")
                    extraction_job = extraction_job[0]
                
                # Build job dict with BOTH states
                # extraction_state is primary for business logic
                # queue_state reflects job runner status
                job_dict = {
                    'id': extraction_job.id,
                    'queue_job_id': queue_job.id,  # Queue job ID for opening queue.job form
                    'name': extraction_job.name,
                    'document_type': extraction_job.document_type,
                    'extraction_state': extraction_job.state,  # NEW: Business logic state (primary)
                    'queue_state': queue_job.state,           # Job runner state (secondary)
                    'current_step': extraction_job.current_step,  # NEW: Current step for progress stepper
                    'create_date': queue_job.date_created.isoformat() if queue_job.date_created else '',  # ISO format for timezone handling
                    'uuid': extraction_job.uuid,  # For bus subscription in progress-only mode
                    'merged_pdf_url': f'/web/content/{extraction_job.attachment_id.id}',  # PDF preview URL
                }

                # Include progress based on extraction.job state (more accurate)
                # Show progress if extraction is processing OR queue is started
                if extraction_job.state == 'processing' or queue_job.state == 'started':
                    job_dict['progress'] = extraction_job.progress or 0
                    job_dict['progress_message'] = extraction_job.progress_message or ''
                else:
                    job_dict['progress'] = 0
                    job_dict['progress_message'] = ''

                # Include result action if extraction succeeded
                # Check extraction_state='done' (primary) AND has result
                if extraction_job.state == 'done' and extraction_job.result_action_json:
                    # Load the stored action
                    action = json.loads(extraction_job.result_action_json)
                    
                    # If we have a linked extraction record, add res_id to open it
                    if extraction_job.extraction_id:
                        action['res_id'] = extraction_job.extraction_id.id
                        # Remove context since we're opening a specific record
                        action.pop('context', None)
                    
                    job_dict['result_action_json'] = json.dumps(action, ensure_ascii=False)
                
                # Include error message if extraction failed
                # Check extraction_state='error' (primary) OR queue_state='failed' (fallback)
                if extraction_job.state == 'error' or queue_job.state == 'failed':
                    # Prefer extraction.job error (more detailed), fallback to queue.job
                    if extraction_job.error_message:
                        job_dict['error_message'] = extraction_job.error_message
                    elif queue_job.exc_message:
                        job_dict['error_message'] = queue_job.exc_message
                    else:
                        job_dict['error_message'] = 'Unknown error'
                
                jobs_list.append(job_dict)
            
            return {
                'jobs': jobs_list,
                'total': total_count,
                'has_more': (offset + limit) < total_count
            }
            
        except Exception as e:
            _logger.exception("Failed to get extraction jobs")
            return {
                'jobs': [],
                'total': 0,
                'has_more': False
            }

    
    @http.route('/document_extractor/substance_dashboard_data', type='json', auth='user', methods=['POST'])
    def get_substance_dashboard_data(self, substance_id=None, organization_id=None, year_from=None, year_to=None):
        """
        Get aggregated data for substance dashboard

        Args:
            substance_id (int): Controlled substance ID
            organization_id (int, optional): Filter by organization
            year_from (int, optional): Start year
            year_to (int, optional): End year

        Returns:
            dict: Dashboard data with KPIs, charts, and details
        """
        try:
            # Validate substance_id
            if not substance_id:
                return {
                    'error': True,
                    'message': 'Substance ID is required'
                }

            # Call substance.aggregate model method
            SubstanceAggregate = request.env['substance.aggregate'].sudo()
            dashboard_data = SubstanceAggregate.get_dashboard_data(
                substance_id=substance_id,
                organization_id=organization_id,
                year_from=year_from,
                year_to=year_to
            )

            return dashboard_data

        except Exception as e:
            _logger.error(f'Error fetching substance dashboard data: {str(e)}', exc_info=True)
            return {
                'error': True,
                'message': str(e)
            }

    @http.route('/document_extractor/company_dashboard_data', type='json', auth='user', methods=['POST'])
    def get_company_dashboard_data(self, organization_id=None, year_from=None, year_to=None):
        """
        Get aggregated data for company dashboard

        Args:
            organization_id (int): Organization/Partner ID (required)
            year_from (int, optional): Start year filter
            year_to (int, optional): End year filter

        Returns:
            dict: Dashboard data with company info, KPIs, charts, and tab data
        """
        try:
            if not organization_id:
                return {'error': True, 'message': 'Organization ID is required'}

            Partner = request.env['res.partner'].sudo()
            Document = request.env['document.extraction'].sudo()

            # Get organization info
            organization = Partner.browse(organization_id)
            if not organization.exists():
                return {'error': True, 'message': 'Organization not found'}

            # Build domain for documents
            domain = [('organization_id', '=', organization_id)]
            if year_from:
                domain.append(('year', '>=', year_from))
            if year_to:
                domain.append(('year', '<=', year_to))

            documents = Document.search(domain)

            # Eager load all related One2many fields to prevent N+1 queries
            # Force ORM to fetch all relations in batch
            documents.mapped('substance_usage_ids.substance_id')
            documents.mapped('substance_usage_ids.substance_name')
            documents.mapped('substance_usage_ids.usage_type')
            documents.mapped('substance_usage_ids.avg_quantity_kg')
            documents.mapped('substance_usage_ids.avg_quantity_co2')
            documents.mapped('substance_usage_ids.year_1_quantity_kg')
            documents.mapped('substance_usage_ids.year_2_quantity_kg')
            documents.mapped('substance_usage_ids.year_3_quantity_kg')
            documents.mapped('substance_usage_ids.year_1_quantity_co2')
            documents.mapped('substance_usage_ids.year_2_quantity_co2')
            documents.mapped('substance_usage_ids.year_3_quantity_co2')
            documents.mapped('quota_usage_ids.substance_name')
            documents.mapped('quota_usage_ids.total_quota_kg')
            documents.mapped('equipment_product_ids.product_type')
            documents.mapped('equipment_product_ids.equipment_type_id')
            documents.mapped('equipment_ownership_ids.equipment_type_id')
            documents.mapped('collection_recycling_ids.activity_type')
            documents.mapped('collection_recycling_ids.quantity_kg')
            documents.mapped('collection_recycling_report_ids.collection_quantity_kg')
            documents.mapped('equipment_product_report_ids')
            documents.mapped('equipment_ownership_report_ids')

            # Aggregate KPIs
            all_substance_usage = documents.mapped('substance_usage_ids').filtered(lambda r: not r.is_title)
            all_quota_usage = documents.mapped('quota_usage_ids')

            total_substances = len(set(
                list(all_substance_usage.mapped('substance_name')) +
                list(all_quota_usage.mapped('substance_name'))
            ))

            # Calculate total kg and CO2e from substance usage
            # FIX: Use correct average calculation (divide by count of non-null years)
            total_kg = 0
            total_co2e = 0
            for usage in all_substance_usage:
                kg, co2 = self._calculate_avg_quantity(usage)
                total_kg += kg
                total_co2e += co2

            # Recovery rate calculation
            # Form 01: collection.recycling with activity_type='collection'
            total_collected_form01 = 0
            for doc in documents.filtered(lambda d: d.document_type == '01'):
                collection_recs = doc.collection_recycling_ids.filtered(lambda r: r.activity_type == 'collection')
                total_collected_form01 += sum(collection_recs.mapped('quantity_kg'))

            # Form 02: collection.recycling.report (field: collection_quantity_kg)
            total_collected_form02 = sum(
                documents.filtered(lambda d: d.document_type == '02')
                .mapped('collection_recycling_report_ids.collection_quantity_kg')
            )
            total_collected = total_collected_form01 + total_collected_form02
            recovery_rate = (total_collected / total_kg * 100) if total_kg > 0 else 0

            # Trend data by year and substance
            trend_data = {}
            for doc in documents:
                year = doc.year
                if year not in trend_data:
                    trend_data[year] = {}

                # Aggregate by substance
                for usage in doc.substance_usage_ids.filtered(lambda r: not r.is_title):
                    substance = usage.substance_name
                    if substance not in trend_data[year]:
                        trend_data[year][substance] = 0

                    # FIX: Use correct average calculation
                    kg, _ = self._calculate_avg_quantity(usage)
                    trend_data[year][substance] += kg

            # Quota allocated vs used
            quota_data = []
            for doc in documents.filtered(lambda d: d.document_type == '02'):
                for quota in doc.quota_usage_ids.filtered(lambda q: not q.is_title):
                    quota_data.append({
                        'year': doc.year,
                        'quota_allocated': quota.allocated_quota_kg or 0,
                        'quota_used': quota.total_quota_kg or 0,
                        'substance_name': quota.substance_name
                    })

            # Get unique activity fields
            all_activity_fields = set()
            for doc in documents:
                all_activity_fields.update(doc.activity_field_ids.mapped('name'))

            return {
                'error': False,
                'company_info': {
                    'name': organization.name,
                    'business_id': organization.business_id or '',
                    'street': organization.street or '',
                    'city': organization.city or '',
                    'country': organization.country_id.name if organization.country_id else '',
                    'activity_fields': ', '.join(list(all_activity_fields)) if all_activity_fields else 'N/A',
                },
                'kpis': {
                    'total_substances': total_substances or 0,
                    'total_kg': total_kg or 0,
                    'total_co2e': total_co2e or 0,
                    'recovery_rate': recovery_rate or 0,
                },
                'charts': {
                    'trend_by_year': trend_data or {},
                    'quota_data': quota_data or [],
                },
                'tabs': {
                    'table_1_1': self._get_table_1_1_data(documents),
                    'table_1_2': self._get_table_1_2_data(documents),
                    'table_1_3': self._get_table_1_3_data(documents),
                    'table_2_1': self._get_table_2_1_data(documents),
                    'table_2_4': self._get_table_2_4_data(documents),
                    'ocr_history': self._get_ocr_history_data(documents),
                }
            }

        except Exception as e:
            _logger.error(f'Error fetching company dashboard data: {str(e)}', exc_info=True)
            return {'error': True, 'message': str(e)}

    def _get_table_1_1_data(self, documents):
        """Get Table 1.1 data (Production/Import/Export)"""
        records = []
        for doc in documents.filtered(lambda d: d.has_table_1_1):
            for usage in doc.substance_usage_ids.filtered(lambda r: not r.is_title):
                # FIX: Calculate correct average
                kg, co2 = self._calculate_avg_quantity(usage)

                records.append({
                    'year': doc.year,
                    'activity': dict(usage._fields['usage_type'].selection).get(usage.usage_type, ''),
                    'substance_name': usage.substance_name,
                    'quantity_kg': kg,
                    'co2e': co2,
                })
        return records

    def _get_table_1_2_data(self, documents):
        """Get Table 1.2 data (Equipment containing substances)"""
        records = []
        for doc in documents.filtered(lambda d: d.has_table_1_2):
            for equipment in doc.equipment_product_ids:
                records.append({
                    'year': doc.year,
                    'equipment_type': equipment.equipment_type,
                    'substance_name': equipment.substance_name,
                    'quantity': equipment.quantity or 0,
                    'capacity': equipment.capacity or 0,
                })
        return records

    def _get_table_1_3_data(self, documents):
        """Get Table 1.3 data (Equipment ownership)"""
        records = []
        for doc in documents.filtered(lambda d: d.has_table_1_3):
            for equipment in doc.equipment_ownership_ids:
                records.append({
                    'year': doc.year,
                    'equipment_type': equipment.equipment_type,
                    'year_start': equipment.start_year or '',
                    'capacity': equipment.capacity or '',
                    'quantity': equipment.equipment_quantity or 0,
                    'substance_name': equipment.substance_name,
                    'refill_amount': equipment.substance_quantity_per_refill or '',
                    'refill_frequency': equipment.refill_frequency or '',
                })
        return records

    def _get_table_2_1_data(self, documents):
        """Get Table 2.1 data (Quota usage)"""
        records = []
        for doc in documents.filtered(lambda d: d.document_type == '02' and d.has_table_2_1):
            for quota in doc.quota_usage_ids.filtered(lambda q: not q.is_title):
                records.append({
                    'year': doc.year,
                    'substance_name': quota.substance_name,
                    'quota_allocated_kg': quota.allocated_quota_kg or 0,
                    'quota_used_kg': quota.total_quota_kg or 0,
                    'hs_code': quota.hs_code or '',
                })
        return records

    def _get_table_2_4_data(self, documents):
        """Get Table 2.4 data (Collection & Recycling)"""
        records = []
        # Form 01 - collection.recycling uses activity_type field
        for doc in documents.filtered(lambda d: d.document_type == '01' and d.has_table_1_4):
            # Group by substance
            substance_data = {}
            for record in doc.collection_recycling_ids:
                substance = record.substance_name
                if substance not in substance_data:
                    substance_data[substance] = {
                        'year': doc.year,
                        'substance_name': substance,
                        'collected': 0,
                        'reused': 0,
                        'recycled': 0,
                        'destroyed': 0,
                    }
                # Add based on activity_type
                if record.activity_type == 'collection':
                    substance_data[substance]['collected'] += record.quantity_kg or 0
                elif record.activity_type == 'reuse':
                    substance_data[substance]['reused'] += record.quantity_kg or 0
                elif record.activity_type == 'recycle':
                    substance_data[substance]['recycled'] += record.quantity_kg or 0
                elif record.activity_type == 'disposal':
                    substance_data[substance]['destroyed'] += record.quantity_kg or 0
            records.extend(substance_data.values())

        # Form 02 - collection.recycling.report has separate fields
        for doc in documents.filtered(lambda d: d.document_type == '02' and d.has_table_2_4):
            for report in doc.collection_recycling_report_ids:
                records.append({
                    'year': doc.year,
                    'substance_name': report.substance_name,
                    'collected': report.collection_quantity_kg or 0,
                    'reused': report.reuse_quantity_kg or 0,
                    'recycled': report.recycle_quantity_kg or 0,
                    'destroyed': report.disposal_quantity_kg or 0,
                })
        return records

    def _get_ocr_history_data(self, documents):
        """Get OCR extraction history"""
        records = []
        for doc in documents:
            records.append({
                'year': doc.year,
                'document_type': 'Mẫu 01' if doc.document_type == '01' else 'Mẫu 02',
                'pdf_filename': doc.pdf_filename,
                'create_date': doc.create_date.strftime('%Y-%m-%d %H:%M') if doc.create_date else '',
                'create_uid': doc.create_uid.name if doc.create_uid else '',
            })
        return records

    @http.route('/document_extractor/equipment_dashboard_data', type='json', auth='user', methods=['POST'])
    def get_equipment_dashboard_data(self, equipment_type_id=None, year_from=None, year_to=None):
        """
        Get aggregated data for equipment type dashboard

        Args:
            equipment_type_id (int): Equipment type ID (required)
            year_from (int, optional): Start year filter
            year_to (int, optional): End year filter

        Returns:
            dict: Dashboard data with equipment info, KPIs, and charts
        """
        try:
            if not equipment_type_id:
                return {'error': True, 'message': 'Equipment type ID is required'}

            EquipmentType = request.env['equipment.type'].sudo()
            EquipmentProduct = request.env['equipment.product'].sudo()
            EquipmentOwnership = request.env['equipment.ownership'].sudo()

            # Get equipment type info
            equipment_type = EquipmentType.browse(equipment_type_id)
            if not equipment_type.exists():
                return {'error': True, 'message': 'Equipment type not found'}

            # Build domain - use equipment_type_id (Many2one to equipment.type)
            domain = [('equipment_type_id', '=', equipment_type_id)]
            if year_from:
                domain.append(('document_id.year', '>=', year_from))
            if year_to:
                domain.append(('document_id.year', '<=', year_to))

            # Get equipment records from both models
            equipment_products = EquipmentProduct.search(domain)
            equipment_ownerships = EquipmentOwnership.search(domain)

            # Eager load relations to prevent N+1 queries
            equipment_products.mapped('document_id.year')
            equipment_products.mapped('document_id.organization_id')
            equipment_products.mapped('substance_id.name')
            equipment_products.mapped('substance_id.gwp')
            equipment_ownerships.mapped('document_id.year')
            equipment_ownerships.mapped('document_id.organization_id')
            equipment_ownerships.mapped('substance_name')

            # Aggregate KPIs
            total_count = len(equipment_products) + len(equipment_ownerships)

            # Only count equipment.product
            total_kg = 0  # substance_quantity_per_unit is now Char (2025-12-18)
            # for eq in equipment_products:
            #     total_kg += (eq.substance_quantity_per_unit or 0) * (eq.quantity or 0)
            # equipment.ownership calculation removed (refill fields are now Char)

            # Capacity is Char field, try to parse or skip
            total_capacity = 0  # Skip for now as it's Char type
            # avg_refill_freq removed (refill_frequency is now Char)

            # Pre-fetch all substances to avoid N+1 queries
            all_substance_names = set()
            all_substance_names.update(equipment_products.mapped('substance_name'))
            all_substance_names.update(equipment_ownerships.mapped('substance_name'))
            # Remove None/empty strings
            all_substance_names = {name for name in all_substance_names if name}

            substances = request.env['controlled.substance'].sudo().search([
                ('name', 'in', list(all_substance_names))
            ])
            # Build lookup dictionary
            gwp_by_name = {s.name: s.gwp for s in substances}

            # GWP calculation
            total_co2e = 0  # substance_quantity_per_unit is now Char (2025-12-18)
            # for eq in equipment_products:
            #     gwp = gwp_by_name.get(eq.substance_name, 0)
            #     kg = (eq.substance_quantity_per_unit or 0) * (eq.quantity or 0)
            #     total_co2e += kg * gwp / 1000  # Convert to tons
            # equipment.ownership CO2e calculation removed (refill fields are now Char)

            # Charts data
            # Trend by year
            trend_by_year = {}
            for eq in equipment_products:
                year = eq.document_id.year
                if year not in trend_by_year:
                    trend_by_year[year] = 0
                trend_by_year[year] += 1
            for eq in equipment_ownerships:
                year = eq.document_id.year
                if year not in trend_by_year:
                    trend_by_year[year] = 0
                trend_by_year[year] += 1

            # By substance
            by_substance = {}  # substance_quantity_per_unit is now Char (2025-12-18)
            # for eq in equipment_products:
            #     substance = eq.substance_name
            #     if substance not in by_substance:
            #         by_substance[substance] = 0
            #     kg = (eq.substance_quantity_per_unit or 0) * (eq.quantity or 0)
            #     by_substance[substance] += kg
            # equipment.ownership removed from by_substance (refill fields are now Char)

            # By company
            by_company = {}
            for eq in equipment_products:
                org = eq.document_id.organization_id.name
                if org not in by_company:
                    by_company[org] = {'capacity': 0, 'count': 0}
                by_company[org]['count'] += 1
            for eq in equipment_ownerships:
                org = eq.document_id.organization_id.name
                if org not in by_company:
                    by_company[org] = {'capacity': 0, 'count': 0}
                by_company[org]['count'] += 1

            # Equipment details
            details = []
            for eq in equipment_ownerships:
                details.append({
                    'organization_name': eq.document_id.organization_id.name,
                    'year_start': eq.start_year or '',
                    'capacity': eq.capacity or '',
                    'quantity': eq.equipment_quantity or 0,
                    'substance_name': eq.substance_name,
                    'refill_amount': eq.substance_quantity_per_refill or '',
                    'refill_frequency': eq.refill_frequency or '',
                })

            # Get unique companies and substances
            unique_companies = set()
            unique_substances = set()
            for eq in equipment_products:
                if eq.document_id.organization_id:
                    unique_companies.add(eq.document_id.organization_id.id)
                if eq.substance_name:
                    unique_substances.add(eq.substance_name)
            for eq in equipment_ownerships:
                if eq.document_id.organization_id:
                    unique_companies.add(eq.document_id.organization_id.id)
                if eq.substance_name:
                    unique_substances.add(eq.substance_name)

            return {
                'error': False,
                'equipment_info': {
                    'name': equipment_type.name,
                    'description': equipment_type.description or '',
                    'capacity_range': f"{equipment_type.min_capacity or 0} - {equipment_type.max_capacity or 0} kW",
                    'total_companies': len(unique_companies),
                    'common_substances': ', '.join(list(unique_substances)[:5]) if unique_substances else 'N/A',
                    'total_capacity': total_capacity,
                },
                'kpis': {
                    'total_count': total_count or 0,
                    'total_kg': total_kg or 0,
                    'total_co2e': total_co2e or 0,
                    # avg_refill_frequency removed
                },
                'charts': {
                    'trend_by_year': [{'year': k, 'count': v} for k, v in sorted(trend_by_year.items())],
                    'by_substance': [{'substance': k, 'total_kg': v} for k, v in by_substance.items()],
                    'by_company': [{'company': k, 'capacity': v['capacity'], 'count': v['count']} for k, v in by_company.items()],
                },
                'details': details[:100],  # Limit to 100 records
            }

        except Exception as e:
            _logger.error(f'Error fetching equipment dashboard data: {str(e)}', exc_info=True)
            return {'error': True, 'message': str(e)}

    @http.route('/document_extractor/hfc_dashboard_data', type='json', auth='user', methods=['POST'])
    def get_hfc_dashboard_data(self, filters=None, **kwargs):
        """
        HFC Dashboard data endpoint (REBUILT to reuse export logic)

        Uses _collect_export_data() for data collection, then aggregates
        in Python for KPIs and charts.

        Args:
            filters (dict): Filter criteria from frontend (13 filters)

        Returns:
            dict: {
                error, kpis, charts: {trend_by_year_substance, by_activity_type, top_10_records, pivot_data},
                filter_metadata: {available_years}
            }
        """
        try:
            filters = filters or {}
            _logger.info(f"HFC Dashboard: Received filters: {filters}")

            # Step 1: Collect filtered data (REUSE FROM EXPORT)
            data = self._collect_export_data(filters)
            _logger.info(f"HFC Dashboard: {len(data['documents'])} documents collected")

            # Step 2: Aggregate KPIs
            kpis = self._aggregate_dashboard_kpis(data)

            # Step 3: Aggregate charts
            trend_data = self._aggregate_trend_by_year_substance(data)
            activity_data = self._aggregate_by_activity_type(data)
            top_10 = self._aggregate_top_10_records(data)
            pivot = self._aggregate_pivot_data(data)

            # Step 4: Extract available years for filter metadata
            available_years = sorted(set(data['documents'].mapped('year'))) if data['documents'] else []

            # Step 5: Return response
            return {
                'error': False,
                'kpis': kpis,
                'charts': {
                    'trend_by_year_substance': trend_data,
                    'by_activity_type': activity_data,
                    'top_10_records': top_10,
                    'pivot_data': pivot,
                },
                'filter_metadata': {
                    'available_years': available_years
                }
            }

        except Exception as e:
            _logger.error(f"HFC Dashboard error: {str(e)}", exc_info=True)
            return {
                'error': True,
                'message': f'Lỗi tải dữ liệu dashboard: {str(e)}'
            }

    @http.route('/document_extractor/recovery_dashboard_data', type='json', auth='user', methods=['POST'])
    def get_recovery_dashboard_data(self, substance_id=None, organization_id=None, year_from=None, year_to=None):
        """
        Get aggregated data for recovery dashboard

        Args:
            substance_id (int, optional): Filter by substance
            organization_id (int, optional): Filter by organization
            year_from (int, optional): Start year filter
            year_to (int, optional): End year filter

        Returns:
            dict: Dashboard data with recovery KPIs and charts
        """
        try:
            Document = request.env['document.extraction'].sudo()

            # Build domain
            domain = []
            if substance_id:
                substance_name = request.env['controlled.substance'].sudo().browse(substance_id).name
                domain.append('|')
                domain.append(('collection_recycling_ids.substance_name', '=', substance_name))
                domain.append(('collection_recycling_report_ids.substance_name', '=', substance_name))
            if organization_id:
                domain.append(('organization_id', '=', organization_id))
            if year_from:
                domain.append(('year', '>=', year_from))
            if year_to:
                domain.append(('year', '<=', year_to))

            documents = Document.search(domain)

            # Eager load relations to prevent N+1 queries
            documents.mapped('collection_recycling_ids.activity_type')
            documents.mapped('collection_recycling_ids.quantity_kg')
            documents.mapped('collection_recycling_ids.substance_name')
            documents.mapped('collection_recycling_report_ids.collection_quantity_kg')
            documents.mapped('collection_recycling_report_ids.reuse_quantity_kg')
            documents.mapped('collection_recycling_report_ids.recycle_quantity_kg')
            documents.mapped('collection_recycling_report_ids.disposal_quantity_kg')
            documents.mapped('organization_id.name')

            # Aggregate KPIs
            # Form 01: collection.recycling with activity_type
            total_collected = 0
            total_reused = 0
            total_recycled = 0
            total_destroyed = 0

            for doc in documents.filtered(lambda d: d.document_type == '01'):
                for record in doc.collection_recycling_ids:
                    if record.activity_type == 'collection':
                        total_collected += record.quantity_kg or 0
                    elif record.activity_type == 'reuse':
                        total_reused += record.quantity_kg or 0
                    elif record.activity_type == 'recycle':
                        total_recycled += record.quantity_kg or 0
                    elif record.activity_type == 'disposal':
                        total_destroyed += record.quantity_kg or 0

            # Form 02: collection.recycling.report has separate fields
            for doc in documents.filtered(lambda d: d.document_type == '02'):
                for report in doc.collection_recycling_report_ids:
                    total_collected += report.collection_quantity_kg or 0
                    total_reused += report.reuse_quantity_kg or 0
                    total_recycled += report.recycle_quantity_kg or 0
                    total_destroyed += report.disposal_quantity_kg or 0

            # Charts
            # Trend by year
            trend_by_year = {}
            for doc in documents:
                year = doc.year
                if year not in trend_by_year:
                    trend_by_year[year] = 0

                # Form 01
                if doc.document_type == '01':
                    collected = sum(doc.collection_recycling_ids.filtered(lambda r: r.activity_type == 'collection').mapped('quantity_kg'))
                    trend_by_year[year] += collected
                # Form 02
                else:
                    collected = sum(doc.collection_recycling_report_ids.mapped('collection_quantity_kg'))
                    trend_by_year[year] += collected

            # By substance (reuse)
            reuse_by_substance = {}
            for doc in documents:
                # Form 01
                if doc.document_type == '01':
                    for record in doc.collection_recycling_ids.filtered(lambda r: r.activity_type == 'reuse'):
                        substance = record.substance_name
                        if substance not in reuse_by_substance:
                            reuse_by_substance[substance] = 0
                        reuse_by_substance[substance] += record.quantity_kg or 0
                # Form 02
                else:
                    for report in doc.collection_recycling_report_ids:
                        substance = report.substance_name
                        if substance not in reuse_by_substance:
                            reuse_by_substance[substance] = 0
                        reuse_by_substance[substance] += report.reuse_quantity_kg or 0

            # By technology (recycle)
            recycle_by_technology = {}
            for doc in documents:
                # Form 01 - no technology field in collection.recycling
                # Form 02 - collection.recycling.report has recycle_technology
                if doc.document_type == '02':
                    for report in doc.collection_recycling_report_ids:
                        tech = report.recycle_technology or 'Unknown'
                        if tech not in recycle_by_technology:
                            recycle_by_technology[tech] = 0
                        recycle_by_technology[tech] += report.recycle_quantity_kg or 0

            # Details
            details = []
            for doc in documents:
                # Form 01 - group by substance
                if doc.document_type == '01':
                    substance_details = {}
                    for record in doc.collection_recycling_ids:
                        substance = record.substance_name
                        if substance not in substance_details:
                            substance_details[substance] = {
                                'organization_name': doc.organization_id.name,
                                'substance_name': substance,
                                'collected': 0,
                                'reused': 0,
                                'recycled': 0,
                                'destroyed': 0,
                                'technology': '',
                                'location': '',
                            }
                        if record.activity_type == 'collection':
                            substance_details[substance]['collected'] += record.quantity_kg or 0
                        elif record.activity_type == 'reuse':
                            substance_details[substance]['reused'] += record.quantity_kg or 0
                        elif record.activity_type == 'recycle':
                            substance_details[substance]['recycled'] += record.quantity_kg or 0
                        elif record.activity_type == 'disposal':
                            substance_details[substance]['destroyed'] += record.quantity_kg or 0
                    details.extend(substance_details.values())

                # Form 02
                else:
                    for report in doc.collection_recycling_report_ids:
                        details.append({
                            'organization_name': doc.organization_id.name,
                            'substance_name': report.substance_name,
                            'collected': report.collection_quantity_kg or 0,
                            'reused': report.reuse_quantity_kg or 0,
                            'recycled': report.recycle_quantity_kg or 0,
                            'destroyed': report.disposal_quantity_kg or 0,
                            'technology': report.recycle_technology or '',
                            'location': report.collection_location or '',
                        })

            # Get main substances from both Form 01 and Form 02
            all_substances = set()
            for doc in documents:
                if doc.document_type == '01':
                    all_substances.update(list(doc.collection_recycling_ids.mapped('substance_name')))
                else:
                    all_substances.update(list(doc.collection_recycling_report_ids.mapped('substance_name')))

            # Remove empty strings and take first 5
            main_substances_list = [s for s in all_substances if s][:5]
            main_substances = ', '.join(main_substances_list) if main_substances_list else 'N/A'

            years = documents.mapped('year')
            year_range = f"{min(years)} - {max(years)}" if years else 'N/A'

            return {
                'error': False,
                'info': {
                    'total_companies': len(set(documents.mapped('organization_id'))),
                    'main_substances': main_substances,
                    'year_range': year_range,
                },
                'kpis': {
                    'total_collected': total_collected or 0,
                    'total_reused': total_reused or 0,
                    'total_recycled': total_recycled or 0,
                    'total_destroyed': total_destroyed or 0,
                },
                'charts': {
                    'trend_by_year': [{'year': k, 'collected': v} for k, v in sorted(trend_by_year.items())],
                    'reuse_by_substance': [{'substance': k, 'reused': v} for k, v in reuse_by_substance.items()],
                    'recycle_by_technology': [{'technology': k, 'recycled': v} for k, v in recycle_by_technology.items()],
                },
                'details': details[:100],
            }

        except Exception as e:
            _logger.error(f'Error fetching recovery dashboard data: {str(e)}', exc_info=True)
            return {'error': True, 'message': str(e)}

    # ===== HFC DASHBOARD AGGREGATION HELPERS =====

    def _aggregate_dashboard_kpis(self, data):
        """
        Aggregate KPIs from filtered documents

        Args:
            data (dict): {
                'documents': RecordSet[document.extraction],
                'substance_filter_ids': [int] or None,
                'filters': dict
            }

        Returns:
            dict: {
                'total_organizations': int,
                'total_kg': float,
                'total_co2e': float,
                'verified_percentage': float
            }
        """
        documents = data['documents']
        substance_filter_ids = data['substance_filter_ids']
        filters = data['filters']

        # KPI 1: Total unique organizations
        total_organizations = len(set(documents.mapped('organization_id').ids))

        # KPI 2 & 3: Total kg and CO2e
        total_kg = 0.0
        total_co2e = 0.0
        quantity_min = filters.get('quantity_min')
        quantity_max = filters.get('quantity_max')

        for doc in documents:
            doc_kg = 0.0
            doc_co2e = 0.0

            if doc.document_type == '01':
                # Form 01: substance_usage_ids (single field)
                for usage in doc.substance_usage_ids:
                    if usage.is_title:
                        continue
                    # Apply substance filter
                    if substance_filter_ids and usage.substance_id.id not in substance_filter_ids:
                        continue

                    # Sum all 3 years
                    qty = (usage.year_1_quantity_kg or 0.0) + \
                          (usage.year_2_quantity_kg or 0.0) + \
                          (usage.year_3_quantity_kg or 0.0)

                    co2e = (usage.year_1_quantity_co2 or 0.0) + \
                           (usage.year_2_quantity_co2 or 0.0) + \
                           (usage.year_3_quantity_co2 or 0.0)

                    doc_kg += qty
                    doc_co2e += co2e

            elif doc.document_type == '02':
                # Form 02: quota usage
                for usage in doc.quota_usage_ids:
                    if usage.is_title:
                        continue
                    # Apply substance filter
                    if substance_filter_ids and usage.substance_id.id not in substance_filter_ids:
                        continue

                    qty = usage.total_quota_kg or 0.0
                    co2e = usage.total_quota_co2 or 0.0

                    doc_kg += qty
                    doc_co2e += co2e

            # Apply quantity filters at document level
            if quantity_min is not None and doc_kg < quantity_min:
                continue
            if quantity_max is not None and doc_kg > quantity_max:
                continue

            total_kg += doc_kg
            total_co2e += doc_co2e

        # KPI 4: Verified percentage
        total_docs = len(documents)
        validated_docs = len(documents.filtered(lambda d: d.state == 'validated'))
        verified_percentage = (validated_docs / total_docs * 100.0) if total_docs > 0 else 0.0

        return {
            'total_organizations': total_organizations,
            'total_kg': total_kg,
            'total_co2e': total_co2e,
            'verified_percentage': verified_percentage
        }

    def _aggregate_trend_by_year_substance(self, data):
        """
        Aggregate trend data by year and substance for bar chart

        Args:
            data (dict): {documents, substance_filter_ids, filters}

        Returns:
            list: [
                {year, substance_id, substance_name, total_kg, co2e},
                ...
            ]
            Sorted by year ASC, substance_name ASC
        """
        documents = data['documents']
        substance_filter_ids = data['substance_filter_ids']

        # Group by (year, substance_id, substance_name)
        trend_groups = {}

        for doc in documents:
            year = doc.year

            if doc.document_type == '01':
                # Form 01: substance_usage_ids (single field)
                for usage in doc.substance_usage_ids:
                    if usage.is_title:
                        continue
                    # Apply substance filter
                    if substance_filter_ids and usage.substance_id.id not in substance_filter_ids:
                        continue

                    substance = usage.substance_id
                    key = (year, substance.id, substance.name)

                    if key not in trend_groups:
                        trend_groups[key] = {'total_kg': 0.0, 'co2e': 0.0}

                    # Sum all 3 years
                    qty = (usage.year_1_quantity_kg or 0.0) + \
                          (usage.year_2_quantity_kg or 0.0) + \
                          (usage.year_3_quantity_kg or 0.0)

                    co2e = (usage.year_1_quantity_co2 or 0.0) + \
                           (usage.year_2_quantity_co2 or 0.0) + \
                           (usage.year_3_quantity_co2 or 0.0)

                    trend_groups[key]['total_kg'] += qty
                    trend_groups[key]['co2e'] += co2e

            elif doc.document_type == '02':
                # Form 02: quota usage
                for usage in doc.quota_usage_ids:
                    if usage.is_title:
                        continue
                    # Apply substance filter
                    if substance_filter_ids and usage.substance_id.id not in substance_filter_ids:
                        continue

                    substance = usage.substance_id
                    key = (year, substance.id, substance.name)

                    if key not in trend_groups:
                        trend_groups[key] = {'total_kg': 0.0, 'co2e': 0.0}

                    qty = usage.total_quota_kg or 0.0
                    co2e = usage.total_quota_co2 or 0.0

                    trend_groups[key]['total_kg'] += qty
                    trend_groups[key]['co2e'] += co2e

        # Convert to list and sort
        result = []
        for (year, substance_id, substance_name), values in trend_groups.items():
            result.append({
                'year': year,
                'substance_id': substance_id,
                'substance_name': substance_name,
                'total_kg': values['total_kg'],
                'co2e': values['co2e']
            })

        # Sort by year ASC, then substance_name ASC (handle None values)
        result.sort(key=lambda x: (x['year'] or 0, x['substance_name'] or ''))

        return result

    def _aggregate_by_activity_type(self, data):
        """
        Aggregate by activity type for pie chart

        Args:
            data (dict): {documents, substance_filter_ids, filters}

        Returns:
            list: [
                {activity_label, total_kg},
                ...
            ]
            Sorted by total_kg DESC
        """
        documents = data['documents']
        substance_filter_ids = data['substance_filter_ids']

        # Activity type labels for Form 02
        ACTIVITY_TYPE_LABELS = {
            'production': 'Sản xuất',
            'import': 'Nhập khẩu',
            'export': 'Xuất khẩu',
        }

        # Group by activity label
        activity_groups = {}

        for doc in documents:
            if doc.document_type == '01':
                # Form 01: substance_usage_ids (single field)
                # Activity type determined by is_title pattern (usage_type field)
                for usage in doc.substance_usage_ids:
                    if usage.is_title:
                        continue
                    # Apply substance filter
                    if substance_filter_ids and usage.substance_id.id not in substance_filter_ids:
                        continue

                    # Map usage_type to label
                    label = ACTIVITY_TYPE_LABELS.get(usage.usage_type, usage.usage_type or 'Khác')

                    if label not in activity_groups:
                        activity_groups[label] = 0.0

                    # Sum all 3 years
                    qty = (usage.year_1_quantity_kg or 0.0) + \
                          (usage.year_2_quantity_kg or 0.0) + \
                          (usage.year_3_quantity_kg or 0.0)

                    activity_groups[label] += qty

            elif doc.document_type == '02':
                # Form 02: Use usage_type field
                for usage in doc.quota_usage_ids:
                    if usage.is_title:
                        continue
                    # Apply substance filter
                    if substance_filter_ids and usage.substance_id.id not in substance_filter_ids:
                        continue

                    label = ACTIVITY_TYPE_LABELS.get(usage.usage_type, usage.usage_type)
                    if label not in activity_groups:
                        activity_groups[label] = 0.0

                    qty = usage.total_quota_kg or 0.0
                    activity_groups[label] += qty

        # Convert to list and sort by total_kg DESC
        result = []
        for label, total_kg in activity_groups.items():
            result.append({
                'activity_label': label,
                'total_kg': total_kg
            })

        result.sort(key=lambda x: x['total_kg'], reverse=True)

        return result

    def _aggregate_top_10_records(self, data):
        """
        Aggregate top 10 records by weight for table

        Args:
            data (dict): {documents, substance_filter_ids, filters}

        Returns:
            list: [
                {
                    organization_id, organization_name, substance_id, substance_name,
                    year, total_kg, co2e, activity_tags, status
                },
                ...
            ]
            Top 10 by total_kg DESC
        """
        documents = data['documents']
        substance_filter_ids = data['substance_filter_ids']

        # Group by (org_id, year, substance_id)
        # Store: {key: {'total_kg', 'co2e', 'doc': document_record}}
        top_groups = {}

        STATE_PRIORITY = {'completed': 3, 'validated': 2, 'draft': 1}

        for doc in documents:
            year = doc.year
            org_id = doc.organization_id.id
            org_name = doc.organization_id.name or ''

            if doc.document_type == '01':
                # Form 01: substance_usage_ids (single field)
                for usage in doc.substance_usage_ids:
                    if usage.is_title:
                        continue
                    # Apply substance filter
                    if substance_filter_ids and usage.substance_id.id not in substance_filter_ids:
                        continue

                    substance = usage.substance_id
                    key = (org_id, year, substance.id)

                    if key not in top_groups:
                        top_groups[key] = {
                            'org_id': org_id,
                            'org_name': org_name,
                            'substance_id': substance.id,
                            'substance_name': substance.name,
                            'year': year,
                            'total_kg': 0.0,
                            'co2e': 0.0,
                            'doc': doc  # Track document
                        }

                    # Sum all 3 years
                    qty = (usage.year_1_quantity_kg or 0.0) + \
                          (usage.year_2_quantity_kg or 0.0) + \
                          (usage.year_3_quantity_kg or 0.0)

                    co2e = (usage.year_1_quantity_co2 or 0.0) + \
                           (usage.year_2_quantity_co2 or 0.0) + \
                           (usage.year_3_quantity_co2 or 0.0)

                    top_groups[key]['total_kg'] += qty
                    top_groups[key]['co2e'] += co2e

                    # Update doc if higher state priority
                    current_doc = top_groups[key]['doc']
                    if STATE_PRIORITY.get(doc.state, 0) > STATE_PRIORITY.get(current_doc.state, 0):
                        top_groups[key]['doc'] = doc

            elif doc.document_type == '02':
                # Form 02: quota usage
                for usage in doc.quota_usage_ids:
                    if usage.is_title:
                        continue
                    # Apply substance filter
                    if substance_filter_ids and usage.substance_id.id not in substance_filter_ids:
                        continue

                    substance = usage.substance_id
                    key = (org_id, year, substance.id)

                    if key not in top_groups:
                        top_groups[key] = {
                            'org_id': org_id,
                            'org_name': org_name,
                            'substance_id': substance.id,
                            'substance_name': substance.name,
                            'year': year,
                            'total_kg': 0.0,
                            'co2e': 0.0,
                            'doc': doc
                        }

                    qty = usage.total_quota_kg or 0.0
                    co2e = usage.total_quota_co2 or 0.0

                    top_groups[key]['total_kg'] += qty
                    top_groups[key]['co2e'] += co2e

                    # Update doc if higher state priority
                    current_doc = top_groups[key]['doc']
                    if STATE_PRIORITY.get(doc.state, 0) > STATE_PRIORITY.get(current_doc.state, 0):
                        top_groups[key]['doc'] = doc

        # Convert to list, extract activity_tags and status, sort, take top 10
        result = []
        for group in top_groups.values():
            doc = group['doc']
            result.append({
                'organization_id': group['org_id'],
                'organization_name': group['org_name'],
                'substance_id': group['substance_id'],
                'substance_name': group['substance_name'],
                'year': group['year'],
                'total_kg': group['total_kg'],
                'co2e': group['co2e'],
                'activity_tags': doc.activity_field_ids.mapped('name'),
                'status': doc.state
            })

        # Sort by total_kg DESC
        result.sort(key=lambda x: x['total_kg'], reverse=True)

        # Take top 10
        return result[:10]

    def _aggregate_pivot_data(self, data):
        """
        Aggregate pivot table data: Organization × Substance × Year

        Args:
            data (dict): {documents, substance_filter_ids, filters}

        Returns:
            list: [
                {
                    organization_id, organization_name, substance_id, substance_name,
                    year_2021_kg, year_2022_kg, ..., total_co2e
                },
                ...
            ]
            Max 50 records, sorted by org_name, substance_name
            Year columns are dynamic based on available years
        """
        documents = data['documents']
        substance_filter_ids = data['substance_filter_ids']

        # Detect available years
        available_years = sorted(set(documents.mapped('year'))) if documents else []

        # Group by (org_id, substance_id)
        # Store: {key: {'org_name', 'substance_name', 'years': {year: kg}, 'total_co2e': float}}
        pivot_groups = {}

        for doc in documents:
            org_id = doc.organization_id.id
            org_name = doc.organization_id.name or ''
            year = doc.year

            if doc.document_type == '01':
                # Form 01: substance_usage_ids (single field)
                for usage in doc.substance_usage_ids:
                    if usage.is_title:
                        continue
                    # Apply substance filter
                    if substance_filter_ids and usage.substance_id.id not in substance_filter_ids:
                        continue

                    substance = usage.substance_id
                    key = (org_id, substance.id)

                    if key not in pivot_groups:
                        pivot_groups[key] = {
                            'org_id': org_id,
                            'org_name': org_name,
                            'substance_id': substance.id,
                            'substance_name': substance.name,
                            'years': {},  # {year: kg}
                            'total_co2e': 0.0
                        }

                    # Sum all 3 years
                    qty = (usage.year_1_quantity_kg or 0.0) + \
                          (usage.year_2_quantity_kg or 0.0) + \
                          (usage.year_3_quantity_kg or 0.0)

                    co2e = (usage.year_1_quantity_co2 or 0.0) + \
                           (usage.year_2_quantity_co2 or 0.0) + \
                           (usage.year_3_quantity_co2 or 0.0)

                    if year not in pivot_groups[key]['years']:
                        pivot_groups[key]['years'][year] = 0.0
                    pivot_groups[key]['years'][year] += qty
                    pivot_groups[key]['total_co2e'] += co2e

            elif doc.document_type == '02':
                # Form 02: quota usage
                for usage in doc.quota_usage_ids:
                    if usage.is_title:
                        continue
                    # Apply substance filter
                    if substance_filter_ids and usage.substance_id.id not in substance_filter_ids:
                        continue

                    substance = usage.substance_id
                    key = (org_id, substance.id)

                    if key not in pivot_groups:
                        pivot_groups[key] = {
                            'org_id': org_id,
                            'org_name': org_name,
                            'substance_id': substance.id,
                            'substance_name': substance.name,
                            'years': {},
                            'total_co2e': 0.0
                        }

                    qty = usage.total_quota_kg or 0.0
                    co2e = usage.total_quota_co2 or 0.0

                    if year not in pivot_groups[key]['years']:
                        pivot_groups[key]['years'][year] = 0.0
                    pivot_groups[key]['years'][year] += qty
                    pivot_groups[key]['total_co2e'] += co2e

        # Convert to list with dynamic year columns
        result = []
        for group in pivot_groups.values():
            row = {
                'organization_id': group['org_id'],
                'organization_name': group['org_name'],
                'substance_id': group['substance_id'],
                'substance_name': group['substance_name'],
                'total_co2e': group['total_co2e']
            }

            # Add year columns
            for year in available_years:
                col_name = f'year_{year}_kg'
                row[col_name] = group['years'].get(year, 0.0)

            result.append(row)

        # Sort by org_name, substance_name (handle None values)
        result.sort(key=lambda x: (x['organization_name'] or '', x['substance_name'] or ''))

        # Limit to 50
        return result[:50]

    @http.route('/document_extractor/export_hfc_report', type='http', auth='user', methods=['POST'], csrf=False)
    def export_hfc_report(self, filters='{}', **kwargs):
        """
        Export HFC dashboard data to Excel using template

        Args:
            filters (str): JSON string of filter criteria (same as get_hfc_dashboard_data)

        Returns:
            HTTP response with Excel file attachment
        """
        import json
        import openpyxl
        from io import BytesIO
        from datetime import datetime
        import os

        try:
            # Parse and validate filters
            try:
                filters_dict = json.loads(filters) if isinstance(filters, str) else filters
            except json.JSONDecodeError as e:
                _logger.error(f"Invalid filter JSON: {str(e)}")
                return request.make_response(
                    json.dumps({'error': True, 'message': 'Định dạng bộ lọc không hợp lệ'}),
                    headers=[('Content-Type', 'application/json')]
                )

            _logger.info(f"Export HFC report with filters: {filters_dict}")

            # Get template path
            module_path = os.path.dirname(os.path.dirname(__file__))
            template_path = os.path.join(module_path, 'static', 'templates', 'HFC_REPORT.xlsx')

            if not os.path.exists(template_path):
                _logger.error(f"Template not found at: {template_path}")
                return request.make_response(
                    json.dumps({'error': True, 'message': 'Template file not found'}),
                    headers=[('Content-Type', 'application/json')]
                )

            # Load template
            try:
                wb = openpyxl.load_workbook(template_path)
            except Exception as e:
                _logger.error(f"Error loading template: {str(e)}", exc_info=True)
                return request.make_response(
                    json.dumps({'error': True, 'message': 'Lỗi khi tải template Excel'}),
                    headers=[('Content-Type', 'application/json')]
                )

            # Collect data using new helper
            try:
                data = self._collect_export_data(filters_dict)
            except Exception as e:
                _logger.error(f"Error collecting data: {str(e)}", exc_info=True)
                return request.make_response(
                    json.dumps({'error': True, 'message': f'Lỗi khi tải dữ liệu: {str(e)}'}),
                    headers=[('Content-Type', 'application/json')]
                )

            # Validate dataset
            if not data['documents']:
                _logger.warning("No documents found matching filters")
                return request.make_response(
                    json.dumps({'error': True, 'message': 'Không tìm thấy dữ liệu phù hợp với bộ lọc. Vui lòng điều chỉnh bộ lọc và thử lại.'}),
                    headers=[('Content-Type', 'application/json')]
                )

            # Fill all 6 sheets
            try:
                self._fill_sheet1_company(wb, data)
                self._fill_sheet2_equipment_ownership(wb, data)
                self._fill_sheet3_equipment_production(wb, data)
                self._fill_sheet4_eol_substances(wb, data)
                self._fill_sheet5_bulk_substances(wb, data)
                self._fill_sheet6_quota_management(wb, data)
            except Exception as e:
                _logger.error(f"Error filling sheets: {str(e)}", exc_info=True)
                return request.make_response(
                    json.dumps({'error': True, 'message': f'Lỗi khi xuất dữ liệu: {str(e)}'}),
                    headers=[('Content-Type', 'application/json')]
                )

            # Save to BytesIO
            try:
                output = BytesIO()
                wb.save(output)
                output.seek(0)
            except Exception as e:
                _logger.error(f"Error saving workbook: {str(e)}", exc_info=True)
                return request.make_response(
                    json.dumps({'error': True, 'message': 'Lỗi khi lưu file Excel'}),
                    headers=[('Content-Type', 'application/json')]
                )

            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'HFC_Report_{timestamp}.xlsx'

            _logger.info(f"Successfully generated {filename} ({len(data['documents'])} documents)")

            # Return as HTTP response
            return request.make_response(
                output.read(),
                headers=[
                    ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                    ('Content-Disposition', f'attachment; filename="{filename}"'),
                ]
            )

        except Exception as e:
            # Catch-all for unexpected errors
            _logger.error(f'Unexpected error in export_hfc_report: {str(e)}', exc_info=True)
            return request.make_response(
                json.dumps({'error': True, 'message': 'Đã xảy ra lỗi không mong muốn. Vui lòng thử lại hoặc liên hệ hỗ trợ.'}),
                headers=[('Content-Type', 'application/json')]
            )

    # ===================================================================================
    # PHASE 1: HELPER UTILITIES (11 methods)
    # ===================================================================================

    # -------------------------------------------------------------------
    # Phase 1.1: Filter & Data Collection Helpers
    # -------------------------------------------------------------------

    def _build_filter_domain(self, filters):
        """
        Build Odoo domain for filtering document.extraction records
        Matches exactly with HFC dashboard filter logic

        Args:
            filters (dict): Filter criteria from frontend
                - status (list): Document states to include
                - year_from (int): Start year filter
                - year_to (int): End year filter
                - activity_field_ids (list): Activity field IDs
                - organization_search (str): Organization name search
                - organization_code (str): Business license number
                - province (str): Province name
                - substance_name (str): Substance name search
                - substance_group_id (int): Substance group ID
                - hs_code (str): HS code search

        Returns:
            list: Odoo domain for document.extraction.search()
        """
        domain = []

        # Base filter: Only completed documents
        if filters.get('status'):
            domain.append(('state', 'in', filters['status']))
        else:
            domain.append(('state', '=', 'completed'))

        # Activity field filter
        if filters.get('activity_field_ids'):
            domain.append(('activity_field_ids', 'in', filters['activity_field_ids']))

        # Year range filter
        if filters.get('year_from'):
            domain.append(('year', '>=', filters['year_from']))
        if filters.get('year_to'):
            domain.append(('year', '<=', filters['year_to']))

        # Organization filters (pre-filter)
        if filters.get('organization_search') or filters.get('organization_code') or filters.get('province'):
            Partner = request.env['res.partner'].sudo()
            org_domain = []

            if filters.get('organization_search'):
                org_domain.append(('name', 'ilike', filters['organization_search']))

            if filters.get('organization_code'):
                org_domain.append(('business_id', 'ilike', filters['organization_code']))

            if filters.get('province'):
                State = request.env['res.country.state'].sudo()
                state_ids = State.search([('name', 'ilike', filters['province'])]).ids
                if state_ids:
                    org_domain.append(('state_id', 'in', state_ids))

            if org_domain:
                org_ids = Partner.search(org_domain).ids
                if org_ids:
                    domain.append(('organization_id', 'in', org_ids))
                else:
                    # No organizations match - return empty domain
                    domain.append(('id', '=', False))

        return domain

    def _collect_export_data(self, filters):
        """
        Centralized data collection for all sheets
        Returns ONLY the latest document per (organization, year, document_type)

        Args:
            filters (dict): Filter criteria from frontend

        Returns:
            dict: {
                'documents': filtered & grouped document.extraction recordset,
                'substance_filter_ids': list of substance IDs to filter (or None),
                'filters': original filters dict
            }
        """
        # Build filter domain
        doc_domain = self._build_filter_domain(filters)

        # Load all documents matching filters
        Document = request.env['document.extraction'].sudo()
        all_documents = Document.search(
            doc_domain,
            order='create_date DESC, id DESC'  # Order by latest first for max() stability
        )

        _logger.info(f"HFC Export: Found {len(all_documents)} total documents matching filters")

        if not all_documents:
            return {
                'documents': Document,
                'substance_filter_ids': None,
                'filters': filters
            }

        # Group by (organization, year, document_type) using Odoo's grouped()
        grouped_docs = all_documents.grouped(
            lambda d: (d.organization_id.id, d.year, d.document_type)
        )

        # Get latest document in each group (by create_date)
        latest_docs = []
        for group_key, docs in grouped_docs.items():
            # docs is already a recordset, sorted by our order clause
            # Take the first one (latest by create_date)
            latest = max(docs, key=lambda d: (d.create_date or d.id))
            latest_docs.append(latest)

        # Sort for consistent output
        latest_docs.sort(
            key=lambda d: (
                d.organization_id.name or '',
                d.year or 0,
                d.document_type or ''
            )
        )

        # Convert back to recordset
        documents = Document.browse([d.id for d in latest_docs])

        _logger.info(f"HFC Export: After grouping - {len(documents)} unique (org, year, type) documents")

        # Build substance filter if needed
        substance_filter_ids = None
        if filters.get('substance_name') or filters.get('substance_group_id') or filters.get('hs_code'):
            Substance = request.env['controlled.substance'].sudo()
            sub_domain = []

            if filters.get('substance_name'):
                sub_domain.append(('name', 'ilike', filters['substance_name']))
            if filters.get('substance_group_id'):
                sub_domain.append(('substance_group_id', '=', filters['substance_group_id']))
            if filters.get('hs_code'):
                sub_domain.append(('hs_code', 'ilike', filters['hs_code']))

            if sub_domain:
                substance_filter_ids = Substance.search(sub_domain).ids
                _logger.info(f"HFC Export: Substance filter active - {len(substance_filter_ids)} substances")

        return {
            'documents': documents,
            'substance_filter_ids': substance_filter_ids,
            'filters': filters
        }

    # -------------------------------------------------------------------
    # Phase 1.2: Generic Excel Writer (Eliminates Duplicate Code)
    # -------------------------------------------------------------------

    def _write_sheet_data(self, workbook, sheet_name, data_rows, start_row=2):
        """
        Generic method to write data rows to Excel sheet with style preservation
        REUSABLE for all 6 sheets - eliminates duplicate cell-writing code

        Args:
            workbook (openpyxl.Workbook): Excel workbook object
            sheet_name (str): Sheet name (e.g., 'HoSo_DoanhNghiep', 'DL_ThietBi_SoHuu')
            data_rows (list of lists): Each inner list = row of cell values
                                       Example: [[1, 'Company A', '123'], [2, 'Company B', '456']]
            start_row (int): Starting row number (default 2, row 1 is header)

        Returns:
            int: Number of rows written

        Features:
            - Writes data values to cells
            - Copies font style from header row (row 1) to maintain consistent formatting

        Example:
            data_rows = [
                [1, 'Company A', '12345', 2024],
                [2, 'Company B', '67890', 2024],
            ]
            self._write_sheet_data(wb, 'HoSo_DoanhNghiep', data_rows)
        """
        from copy import copy
        from openpyxl.styles import Font

        if sheet_name not in workbook.sheetnames:
            _logger.error(f"HFC Export: Sheet '{sheet_name}' not found in template")
            raise ValueError(f"Sheet '{sheet_name}' not found in template")

        sheet = workbook[sheet_name]
        rows_written = 0

        # Get header row (row 1) for style reference
        header_row = 1

        for row_idx, row_data in enumerate(data_rows, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Copy ONLY font name and size from header (no bold/italic/color)
                header_cell = sheet.cell(row=header_row, column=col_idx)
                if header_cell.font:
                    cell.font = Font(
                        name=header_cell.font.name,
                        size=header_cell.font.size
                    )
                if header_cell.alignment:
                    cell.alignment = copy(header_cell.alignment)
                if header_cell.border:
                    cell.border = copy(header_cell.border)

            rows_written += 1

        _logger.info(f"HFC Export: Sheet '{sheet_name}' - {rows_written} rows written")

        return rows_written

    # -------------------------------------------------------------------
    # Phase 1.3: Pattern Resolver Helpers
    # -------------------------------------------------------------------

    def _resolve_is_title_context(self, records, title_field_name='substance_name'):
        """
        Resolve is_title pattern: build mapping from data record → title record

        Pattern explanation (using substance_name as example):
            line 1: is_title=True, substance_name='Sản xuất' (section title)
            line 2: is_title=False, substance_name='R-134a' → maps to 'Sản xuất'
            line 3: is_title=False, substance_name='R-410A' → maps to 'Sản xuất'
            line 4: is_title=True, substance_name='Nhập khẩu' (section title)
            line 5: is_title=False, substance_name='R-22' → maps to 'Nhập khẩu'

        Args:
            records (recordset): Ordered recordset with is_title field
            title_field_name (str): Field to extract from title record
                                   Valid fields: 'substance_name', 'equipment_type', 'product_type'

        Returns:
            dict: {record.id: title_value}
                 Maps each data record (is_title=False) to its title context value

        Example:
            context = self._resolve_is_title_context(substance_usage_ids, 'substance_name')
            # Returns: {2: 'Sản xuất', 3: 'Sản xuất', 5: 'Nhập khẩu'}
        """
        mapping = {}
        current_title_value = None

        # Records should already be ordered by sequence
        for record in records.sorted('sequence'):
            if record.is_title:
                # This is a title row - update context
                current_title_value = record[title_field_name]
            else:
                # This is a data row - map to current title
                if current_title_value is not None:
                    mapping[record.id] = current_title_value

        _logger.debug(f"HFC Export: Resolved {len(mapping)} is_title mappings")

        return mapping

    def _format_capacity(self, record):
        """
        Format capacity field for equipment records
        Logic: If capacity exists → use it, else → combine cooling_capacity/power_capacity with "/"

        Args:
            record: equipment.ownership, equipment.ownership.report,
                   equipment.product, or equipment.product.report record
                   Must have fields: capacity, cooling_capacity, power_capacity

        Returns:
            str or float: Formatted capacity value
                         Examples: "5.5", "2500/1800", ""
        """
        # Use capacity if exists
        if record.capacity:
            return record.capacity

        # Build combined capacity from cooling/power fields
        parts = []
        if record.cooling_capacity:
            parts.append(str(record.cooling_capacity))
        if record.power_capacity:
            parts.append(str(record.power_capacity))

        return '/'.join(parts) if parts else ''

    # -------------------------------------------------------------------
    # Phase 1.4: Unpivot Helpers (Complex Data Transformation)
    # -------------------------------------------------------------------

    def _unpivot_substance_usage(self, usage_record, title_context, doc):
        """
        Unpivot substance_usage record (Form 01):
        1 record → 3 rows (year_1, year_2, year_3)

        Used in Sheet 5 (DL_MoiChat_Bulk) and Sheet 6 (DL_QuanLy_HanNgach)

        Args:
            usage_record (substance.usage): Record with is_title=False
            title_context (dict): Mapping {record.id → usage_type value}
            doc (document.extraction): Parent document

        Returns:
            list of dict: Up to 3 rows (one per year with data)
        """
        usage_type = title_context.get(usage_record.id, 'import')
        usage_type_map = {
            'production': 'Sản xuất',
            'import': 'Nhập khẩu',
            'export': 'Xuất khẩu'
        }

        # Common data for all rows
        base_row = {
            'TenDoanhNghiep': doc.organization_id.name or '',
            'MaSoDN': doc.organization_id.business_id or '',
            'NamBaoCao': doc.year,
            'HoatDong': usage_type_map.get(usage_type, ''),
            'TenChat': usage_record.substance_name or '',
            'GhiChu': usage_record.notes or '',
        }

        rows = []

        # Year 1
        if usage_record.year_1_quantity_kg or usage_record.year_1_quantity_co2:
            rows.append({
                **base_row,
                'NamDuLieu': doc.year_1,
                'Luong_kg': usage_record.year_1_quantity_kg or 0,
                'Luong_CO2td': usage_record.year_1_quantity_co2 or 0,
            })

        # Year 2
        if usage_record.year_2_quantity_kg or usage_record.year_2_quantity_co2:
            rows.append({
                **base_row,
                'NamDuLieu': doc.year_2,
                'Luong_kg': usage_record.year_2_quantity_kg or 0,
                'Luong_CO2td': usage_record.year_2_quantity_co2 or 0,
            })

        # Year 3
        if usage_record.year_3_quantity_kg or usage_record.year_3_quantity_co2:
            rows.append({
                **base_row,
                'NamDuLieu': doc.year_3,
                'Luong_kg': usage_record.year_3_quantity_kg or 0,
                'Luong_CO2td': usage_record.year_3_quantity_co2 or 0,
            })

        return rows

    def _unpivot_quota_usage(self, quota_record, title_context, doc):
        """
        Unpivot quota_usage record (Form 02):
        1 record → 4 rows (allocated, adjusted, used, next_year quotas)

        Used in Sheet 6 (DL_QuanLy_HanNgach) only

        Args:
            quota_record (quota.usage): Record with is_title=False
            title_context (dict): Mapping {record.id → usage_type value}
            doc (document.extraction): Parent document

        Returns:
            list of dict: Up to 4 rows (one per quota type with data)
        """
        usage_type = title_context.get(quota_record.id, 'import')
        usage_type_map = {
            'production': 'Sản xuất',
            'import': 'Nhập khẩu',
            'export': 'Xuất khẩu'
        }

        # Common data for all rows
        base_row = {
            'TenDoanhNghiep': doc.organization_id.name or '',
            'MaSoDN': doc.organization_id.business_id or '',
            'NguonDuLieu': 'Mẫu 02',
            'NamBaoCao': doc.year,
            'NamDuLieu': doc.year,
            'HoatDong': usage_type_map.get(usage_type, ''),
            'TenChat': quota_record.substance_name or '',
            'MaHS': quota_record.hs_code or '',
            'GhiChu': quota_record.notes or ''
        }

        rows = []

        # Row 1: Hạn ngạch được phân bổ
        if quota_record.allocated_quota_kg or quota_record.allocated_quota_co2:
            rows.append({
                **base_row,
                'LoaiThongTin': 'Hạn ngạch được phân bổ',
                'Luong_kg': quota_record.allocated_quota_kg or 0,
                'Luong_CO2td': quota_record.allocated_quota_co2 or 0,
            })

        # Row 2: Hạn ngạch điều chỉnh
        if quota_record.adjusted_quota_kg or quota_record.adjusted_quota_co2:
            rows.append({
                **base_row,
                'LoaiThongTin': 'Hạn ngạch điều chỉnh',
                'Luong_kg': quota_record.adjusted_quota_kg or 0,
                'Luong_CO2td': quota_record.adjusted_quota_co2 or 0,
            })

        # Row 3: Hạn ngạch đã sử dụng
        if quota_record.total_quota_kg or quota_record.total_quota_co2:
            rows.append({
                **base_row,
                'LoaiThongTin': 'Hạn ngạch đã sử dụng',
                'Luong_kg': quota_record.total_quota_kg or 0,
                'Luong_CO2td': quota_record.total_quota_co2 or 0,
            })

        # Row 4: Đăng ký hạn ngạch năm sau
        if quota_record.next_year_quota_kg or quota_record.next_year_quota_co2:
            rows.append({
                **base_row,
                'LoaiThongTin': 'Đăng ký hạn ngạch năm sau',
                'Luong_kg': quota_record.next_year_quota_kg or 0,
                'Luong_CO2td': quota_record.next_year_quota_co2 or 0,
            })

        return rows

    def _unpivot_collection_report(self, report_record, doc):
        """
        Unpivot collection_recycling_report record (Form 02):
        1 record → up to 4 rows (collection, reuse, recycle, disposal)

        Only create rows where quantity_kg > 0

        Used in Sheet 4 (DL_MoiChat_EoL) only

        Args:
            report_record (collection.recycling.report): Record with horizontal structure
            doc (document.extraction): Parent document

        Returns:
            list of dict: 0-4 rows (only activities with quantity > 0)
        """
        # Common data for all rows
        base_row = {
            'TenDoanhNghiep': doc.organization_id.name or '',
            'MaSoDN': doc.organization_id.business_id or '',
            'NamBaoCao': doc.year,
            'NguonDuLieu': 'Mẫu 02 - Bảng 2.4',
            'TenChat': report_record.substance_id.name if report_record.substance_id else '',
        }

        rows = []

        # Row 1: Thu gom
        if report_record.collection_quantity_kg and report_record.collection_quantity_kg > 0:
            rows.append({
                **base_row,
                'HoatDong': 'Thu gom',
                'KhoiLuong_kg': report_record.collection_quantity_kg,
                'ChiTiet_1': report_record.collection_location or '',
                'ChiTiet_2': report_record.storage_location or '',
            })

        # Row 2: Tái sử dụng
        if report_record.reuse_quantity_kg and report_record.reuse_quantity_kg > 0:
            rows.append({
                **base_row,
                'HoatDong': 'Tái sử dụng',
                'KhoiLuong_kg': report_record.reuse_quantity_kg,
                'ChiTiet_1': report_record.reuse_technology or '',
                'ChiTiet_2': '',
            })

        # Row 3: Tái chế
        if report_record.recycle_quantity_kg and report_record.recycle_quantity_kg > 0:
            rows.append({
                **base_row,
                'HoatDong': 'Tái chế',
                'KhoiLuong_kg': report_record.recycle_quantity_kg,
                'ChiTiet_1': report_record.recycle_technology or '',
                'ChiTiet_2': report_record.recycle_usage_location or '',
            })

        # Row 4: Xử lý/Tiêu hủy
        if report_record.disposal_quantity_kg and report_record.disposal_quantity_kg > 0:
            rows.append({
                **base_row,
                'HoatDong': 'Xử lý/Tiêu hủy',
                'KhoiLuong_kg': report_record.disposal_quantity_kg,
                'ChiTiet_1': report_record.disposal_technology or '',
                'ChiTiet_2': report_record.disposal_facility or '',
            })

        return rows

    # ===================================================================================
    # PHASE 2: SHEET FILLER METHODS (6 sheets)
    # ===================================================================================

    def _fill_sheet1_company(self, workbook, data):
        """
        Fill Sheet 1: HoSo_DoanhNghiep (Company/Organization Profile)

        Logic:
            - Receives pre-grouped documents (latest per org/year/type)
            - Extract activity fields as boolean columns

        Args:
            workbook (openpyxl.Workbook): Excel workbook
            data (dict): Data from _collect_export_data() (already grouped)

        Returns:
            None (modifies workbook in-place)
        """
        documents = data['documents']

        data_rows = []

        for idx, doc in enumerate(documents, start=1):
            org = doc.organization_id

            # Get activity field codes for this document
            doc_activity_codes = doc.activity_field_ids.mapped('code')

            # Map document_type to NguonDuLieu
            nguon_du_lieu = 'Mẫu 01' if doc.document_type == '01' else 'Mẫu 02'

            # Build row (20 columns)
            row = [
                idx,                                          # 1. STT
                org.name or '',                               # 2. TenDoanhNghiep
                org.business_id or '',                        # 3. MaSoDN
                doc.year,                                     # 4. NamBaoCao
                nguon_du_lieu,                                # 5. NguonDuLieu
                doc.legal_representative_name or '',          # 6. TenNguoiDaiDienPhapLuat
                doc.legal_representative_position or '',      # 7. ChucVu
                doc.contact_person_name or '',                # 8. TenNguoiDaiDienLienLac
                doc.contact_address or '',                    # 9. DiaChi
                doc.contact_state_id.name if doc.contact_state_id else '',  # 10. Tinh_ThanhPho
                doc.contact_phone or '',                      # 11. DienThoai
                doc.contact_email or '',                      # 12. Email
                # Activity fields (13-20): 'X' if present, '' if not
                'X' if 'production' in doc_activity_codes else '',           # 13. LinhVuc_SanXuatChat
                'X' if 'import' in doc_activity_codes else '',               # 14. LinhVuc_NhapKhauChat
                'X' if 'export' in doc_activity_codes else '',               # 15. LinhVuc_XuatKhauChat
                'X' if 'equipment_production' in doc_activity_codes else '', # 16. LinhVuc_SanXuatThietBi
                'X' if 'equipment_import' in doc_activity_codes else '',     # 17. LinhVuc_NhapKhauThietBi
                'X' if 'ac_ownership' in doc_activity_codes else '',         # 18. LinhVuc_SoHuu_DHKK
                'X' if 'refrigeration_ownership' in doc_activity_codes else '',  # 19. LinhVuc_SoHuu_ThietBiLanh
                'X' if 'collection_recycling' in doc_activity_codes else '', # 20. LinhVuc_ThuGomXuLy
            ]

            data_rows.append(row)

        # Write all rows at once using generic writer
        self._write_sheet_data(workbook, '1_Hoso_DoanhNhiep', data_rows)

    def _fill_sheet5_bulk_substances(self, workbook, data):
        """
        Fill Sheet 5: DL_MoiChat_Bulk (Bulk Substances - Production/Import/Export)

        Logic:
            - Form 01: Unpivot substance_usage (1 record → 3 rows for 3 years)
            - Form 02: Direct mapping from quota_usage (1 record → 1 row)
            - Use is_title pattern to determine HoatDong

        Args:
            workbook (openpyxl.Workbook): Excel workbook
            data (dict): Data from _collect_export_data()

        Returns:
            None (modifies workbook in-place)
        """
        documents = data['documents']
        substance_filter_ids = data['substance_filter_ids']

        data_rows = []

        for doc in documents:
            org = doc.organization_id

            # Form 01: substance_usage_ids (unpivot 3 years)
            if doc.document_type == '01':
                records = doc.substance_usage_ids.sorted('sequence')

                # Resolve is_title context
                title_context = self._resolve_is_title_context(records, 'substance_name')

                for record in records:
                    # Skip title rows
                    if record.is_title:
                        continue

                    # Apply substance filter
                    if substance_filter_ids and record.substance_id.id not in substance_filter_ids:
                        continue

                    # Unpivot: 1 record → up to 3 rows
                    unpivoted_rows = self._unpivot_substance_usage(record, title_context, doc)

                    # Convert dict rows to list rows (9 columns)
                    for row_dict in unpivoted_rows:
                        row = [
                            row_dict['TenDoanhNghiep'],     # 1
                            row_dict['MaSoDN'],             # 2
                            'Mẫu 01 - Bảng 1.1',            # 3. NguonDuLieu
                            row_dict['HoatDong'],           # 4
                            row_dict['TenChat'],            # 5
                            row_dict['NamDuLieu'],          # 6
                            row_dict['Luong_kg'],           # 7
                            row_dict['Luong_CO2td'],        # 8
                            '',                             # 9. MaHS (Form 01 doesn't have)
                        ]
                        data_rows.append(row)

            # Form 02: quota_usage_ids (direct 1-to-1)
            elif doc.document_type == '02':
                records = doc.quota_usage_ids.sorted('sequence')

                # Resolve is_title context
                title_context = self._resolve_is_title_context(records, 'substance_name')

                # Mapping for HoatDong
                usage_type_map = {
                    'production': 'Sản xuất',
                    'import': 'Nhập khẩu',
                    'export': 'Xuất khẩu'
                }

                for record in records:
                    # Skip title rows
                    if record.is_title:
                        continue

                    # Apply substance filter
                    if substance_filter_ids and record.substance_id.id not in substance_filter_ids:
                        continue

                    usage_type = title_context.get(record.id, 'import')

                    row = [
                        org.name or '',                     # 1. TenDoanhNghiep
                        org.business_id or '',              # 2. MaSoDN
                        'Mẫu 02 - Bảng 2.1',                # 3. NguonDuLieu
                        usage_type_map.get(usage_type, ''), # 4. HoatDong
                        record.substance_name or '',        # 5. TenChat
                        doc.year,                           # 6. NamDuLieu
                        record.total_quota_kg or 0,         # 7. Luong_kg
                        record.total_quota_co2 or 0,        # 8. Luong_CO2td
                        record.hs_code or '',               # 9. MaHS
                    ]
                    data_rows.append(row)

        # Write all rows at once
        self._write_sheet_data(workbook, '5_DL_MoiChat_Bulk', data_rows)

    def _fill_sheet2_equipment_ownership(self, workbook, data):
        """
        Fill Sheet 2: DL_ThietBi_SoHuu (Equipment Ownership)

        Logic:
            - Form 01: equipment_ownership_ids
            - Form 02: equipment_ownership_report_ids
            - Use is_title pattern to get PhanLoaiThietBi
            - Format capacity using helper

        Args:
            workbook (openpyxl.Workbook): Excel workbook
            data (dict): Data from _collect_export_data()

        Returns:
            None (modifies workbook in-place)
        """
        documents = data['documents']
        substance_filter_ids = data['substance_filter_ids']

        data_rows = []

        for doc in documents:
            org = doc.organization_id

            # Form 01: equipment_ownership_ids
            if doc.document_type == '01':
                records = doc.equipment_ownership_ids.sorted('sequence')

                # Resolve is_title context for PhanLoaiThietBi
                title_context = self._resolve_is_title_context(records, 'equipment_type')

                for record in records:
                    # Skip title rows
                    if record.is_title:
                        continue

                    # Apply substance filter
                    if substance_filter_ids and record.substance_id and record.substance_id.id not in substance_filter_ids:
                        continue

                    # Get PhanLoaiThietBi from title context
                    phan_loai_thiet_bi = title_context.get(record.id, '')

                    row = [
                        org.name or '',                                  # 1. TenDoanhNghiep
                        org.business_id or '',                           # 2. MaSoDN
                        doc.year,                                        # 3. NamBaoCao
                        'Mẫu 01 - Bảng 1.3',                             # 4. NguonDuLieu
                        phan_loai_thiet_bi,                              # 5. PhanLoaiThietBi
                        record.equipment_type or '',                     # 6. TenLoaiThietBi
                        record.substance_id.name if record.substance_id else '',  # 7. TenChat
                        self._format_capacity(record),                   # 8. NangSuat
                        record.start_year or '',                         # 9. NamSuDung
                        record.equipment_quantity or 0,                  # 10. SoLuong
                        record.refill_frequency or '',                   # 11. TanSuatNapMoi
                        record.substance_quantity_per_refill or 0,       # 12. LuongNapMoi_kg
                        '',                                              # 13. GhiChu_ThietBiMoi (Form 01 doesn't have)
                    ]
                    data_rows.append(row)

            # Form 02: equipment_ownership_report_ids
            elif doc.document_type == '02':
                records = doc.equipment_ownership_report_ids.sorted('sequence')

                # Resolve is_title context
                title_context = self._resolve_is_title_context(records, 'equipment_type')

                for record in records:
                    # Skip title rows
                    if record.is_title:
                        continue

                    # Apply substance filter
                    if substance_filter_ids and record.substance_id and record.substance_id.id not in substance_filter_ids:
                        continue

                    phan_loai_thiet_bi = title_context.get(record.id, '')

                    row = [
                        org.name or '',                                  # 1
                        org.business_id or '',                           # 2
                        doc.year,                                        # 3
                        'Mẫu 02 - Bảng 2.3',                             # 4
                        phan_loai_thiet_bi,                              # 5
                        record.equipment_type or '',                     # 6
                        record.substance_id.name if record.substance_id else '',  # 7
                        self._format_capacity(record),                   # 8
                        record.start_year or '',                         # 9
                        record.equipment_quantity or 0,                  # 10
                        record.refill_frequency or '',                   # 11
                        record.substance_quantity_per_refill or 0,       # 12
                        record.notes or '',                              # 13. GhiChu_ThietBiMoi (Form 02 has this)
                    ]
                    data_rows.append(row)

        # Write all rows
        self._write_sheet_data(workbook, '2_DL_ThietBi_SoHuu', data_rows)

    def _fill_sheet3_equipment_production(self, workbook, data):
        """
        Fill Sheet 3: DL_ThietBi_SX_NK (Equipment Production/Import)

        Logic:
            - Form 01: equipment_product_ids
            - Form 02: equipment_product_report_ids (has production_type field)
            - Similar to Sheet 2 but with HoatDong column

        Args:
            workbook (openpyxl.Workbook): Excel workbook
            data (dict): Data from _collect_export_data()

        Returns:
            None (modifies workbook in-place)
        """
        documents = data['documents']
        substance_filter_ids = data['substance_filter_ids']

        data_rows = []

        for doc in documents:
            org = doc.organization_id

            # Form 01: equipment_product_ids
            if doc.document_type == '01':
                records = doc.equipment_product_ids.sorted('sequence')

                for record in records:
                    # Skip title rows
                    if record.is_title:
                        continue

                    # Apply substance filter
                    if substance_filter_ids and record.substance_id and record.substance_id.id not in substance_filter_ids:
                        continue

                    # Note: Form 01 doesn't have production_type field
                    # Leave HoatDong empty for now (can be inferred later if needed)
                    hoat_dong = ''

                    row = [
                        org.name or '',                                  # 1. TenDoanhNghiep
                        org.business_id or '',                           # 2. MaSoDN
                        doc.year,                                        # 3. NamBaoCao
                        'Mẫu 01 - Bảng 1.2',                             # 4. NguonDuLieu
                        hoat_dong,                                       # 5. HoatDong
                        record.product_type or '',                       # 6. TenLoaiSanPham
                        record.hs_code_id.code if record.hs_code_id else '',  # 7. MaHS
                        self._format_capacity(record),                   # 8. NangSuat
                        record.quantity or 0,                            # 9. SoLuong
                        record.substance_id.name if record.substance_id else '',  # 10. TenChat
                        record.substance_quantity_per_unit or '',        # 11. LuongChatTrongTB_kg
                    ]
                    data_rows.append(row)

            # Form 02: equipment_product_report_ids
            elif doc.document_type == '02':
                records = doc.equipment_product_report_ids.sorted('sequence')

                for record in records:
                    # Skip title rows
                    if record.is_title:
                        continue

                    # Apply substance filter
                    if substance_filter_ids and record.substance_id and record.substance_id.id not in substance_filter_ids:
                        continue

                    # Form 02 has production_type field
                    hoat_dong = 'Sản xuất' if record.production_type == 'production' else 'Nhập khẩu'

                    row = [
                        org.name or '',                                  # 1
                        org.business_id or '',                           # 2
                        doc.year,                                        # 3
                        'Mẫu 02 - Bảng 2.2',                             # 4
                        hoat_dong,                                       # 5
                        record.product_type or '',                       # 6
                        record.hs_code_id.code if record.hs_code_id else '',  # 7
                        self._format_capacity(record),                   # 8
                        record.quantity or 0,                            # 9
                        record.substance_id.name if record.substance_id else '',  # 10
                        record.substance_quantity_per_unit or '',        # 11
                    ]
                    data_rows.append(row)

        # Write all rows
        self._write_sheet_data(workbook, '3_DL_ThietBi_SX_NK', data_rows)

    def _fill_sheet4_eol_substances(self, workbook, data):
        """
        Fill Sheet 4: DL_MoiChat_EoL (End-of-Life Substances - Collection/Recycling)

        Logic:
            - Form 01: collection_recycling_ids (vertical, is_title pattern)
            - Form 02: collection_recycling_report_ids (horizontal → unpivot)
            - Only export rows where quantity_kg > 0

        Args:
            workbook (openpyxl.Workbook): Excel workbook
            data (dict): Data from _collect_export_data()

        Returns:
            None (modifies workbook in-place)
        """
        documents = data['documents']
        substance_filter_ids = data['substance_filter_ids']

        # Activity type mapping
        activity_map = {
            'collection': 'Thu gom',
            'reuse': 'Tái sử dụng',
            'recycle': 'Tái chế',
            'disposal': 'Xử lý/Tiêu hủy'
        }

        data_rows = []

        for doc in documents:
            org = doc.organization_id

            # Form 01: collection_recycling_ids (vertical structure)
            if doc.document_type == '01':
                records = doc.collection_recycling_ids.sorted('sequence')

                # Resolve is_title context
                title_context = self._resolve_is_title_context(records, 'substance_name')

                for record in records:
                    # Skip title rows
                    if record.is_title:
                        continue

                    # Apply substance filter
                    if substance_filter_ids and record.substance_id and record.substance_id.id not in substance_filter_ids:
                        continue

                    # Skip if quantity is 0 or empty
                    if not record.quantity_kg or record.quantity_kg <= 0:
                        continue

                    activity_type = title_context.get(record.id, 'collection')
                    hoat_dong = activity_map.get(activity_type, '')

                    row = [
                        org.name or '',                                  # 1. TenDoanhNghiep
                        org.business_id or '',                           # 2. MaSoDN
                        doc.year,                                        # 3. NamBaoCao
                        'Mẫu 01 - Bảng 1.4',                             # 4. NguonDuLieu
                        record.substance_id.name if record.substance_id else '',  # 5. TenChat
                        hoat_dong,                                       # 6. HoatDong
                        record.quantity_kg or 0,                         # 7. KhoiLuong_kg
                        '',                                              # 8. ChiTiet_1 (Form 01 doesn't have)
                        '',                                              # 9. ChiTiet_2
                    ]
                    data_rows.append(row)

            # Form 02: collection_recycling_report_ids (horizontal → unpivot)
            elif doc.document_type == '02':
                records = doc.collection_recycling_report_ids

                for record in records:
                    # Apply substance filter
                    if substance_filter_ids and record.substance_id and record.substance_id.id not in substance_filter_ids:
                        continue

                    # Unpivot: 1 record → up to 4 rows
                    unpivoted_rows = self._unpivot_collection_report(record, doc)

                    # Convert dict rows to list rows (9 columns)
                    for row_dict in unpivoted_rows:
                        row = [
                            row_dict['TenDoanhNghiep'],     # 1
                            row_dict['MaSoDN'],             # 2
                            row_dict['NamBaoCao'],          # 3
                            row_dict['NguonDuLieu'],        # 4
                            row_dict['TenChat'],            # 5
                            row_dict['HoatDong'],           # 6
                            row_dict['KhoiLuong_kg'],       # 7
                            row_dict['ChiTiet_1'],          # 8
                            row_dict['ChiTiet_2'],          # 9
                        ]
                        data_rows.append(row)

        # Write all rows
        self._write_sheet_data(workbook, '4_DL_MoiChat_EoL', data_rows)

    def _fill_sheet6_quota_management(self, workbook, data):
        """
        Fill Sheet 6: DL_QuanLy_HanNgach (Quota Management) - MOST COMPLEX

        Logic:
            - Form 01: substance_usage_ids (unpivot 3 years, LoaiThongTin = "Lượng sử dụng quá khứ")
            - Form 02: quota_usage_ids (unpivot 4 quota types)

        Args:
            workbook (openpyxl.Workbook): Excel workbook
            data (dict): Data from _collect_export_data()

        Returns:
            None (modifies workbook in-place)
        """
        documents = data['documents']
        substance_filter_ids = data['substance_filter_ids']

        data_rows = []

        for doc in documents:
            org = doc.organization_id

            # Form 01: substance_usage_ids (unpivot 3 years)
            if doc.document_type == '01':
                records = doc.substance_usage_ids.sorted('sequence')

                # Resolve is_title context
                title_context = self._resolve_is_title_context(records, 'substance_name')

                for record in records:
                    # Skip title rows
                    if record.is_title:
                        continue

                    # Apply substance filter
                    if substance_filter_ids and record.substance_id.id not in substance_filter_ids:
                        continue

                    # Unpivot: 1 record → up to 3 rows
                    unpivoted_rows = self._unpivot_substance_usage(record, title_context, doc)

                    # Convert dict rows to list rows (12 columns)
                    for row_dict in unpivoted_rows:
                        row = [
                            row_dict['TenDoanhNghiep'],     # 1
                            row_dict['MaSoDN'],             # 2
                            'Mẫu 01',                       # 3. NguonDuLieu
                            row_dict['NamBaoCao'],          # 4
                            row_dict['NamDuLieu'],          # 5
                            row_dict['HoatDong'],           # 6
                            row_dict['TenChat'],            # 7
                            '',                             # 8. MaHS (Form 01 doesn't have)
                            'Lượng sử dụng quá khứ',        # 9. LoaiThongTin
                            row_dict['Luong_kg'],           # 10
                            row_dict['Luong_CO2td'],        # 11
                            row_dict['GhiChu'],             # 12
                        ]
                        data_rows.append(row)

            # Form 02: quota_usage_ids (unpivot 4 quota types)
            elif doc.document_type == '02':
                records = doc.quota_usage_ids.sorted('sequence')

                # Resolve is_title context
                title_context = self._resolve_is_title_context(records, 'substance_name')

                for record in records:
                    # Skip title rows
                    if record.is_title:
                        continue

                    # Apply substance filter
                    if substance_filter_ids and record.substance_id.id not in substance_filter_ids:
                        continue

                    # Unpivot: 1 record → up to 4 rows
                    unpivoted_rows = self._unpivot_quota_usage(record, title_context, doc)

                    # Convert dict rows to list rows (12 columns)
                    for row_dict in unpivoted_rows:
                        row = [
                            row_dict['TenDoanhNghiep'],     # 1
                            row_dict['MaSoDN'],             # 2
                            row_dict['NguonDuLieu'],        # 3
                            row_dict['NamBaoCao'],          # 4
                            row_dict['NamDuLieu'],          # 5
                            row_dict['HoatDong'],           # 6
                            row_dict['TenChat'],            # 7
                            row_dict['MaHS'],               # 8
                            row_dict['LoaiThongTin'],       # 9
                            row_dict['Luong_kg'],           # 10
                            row_dict['Luong_CO2td'],        # 11
                            row_dict['GhiChu'],             # 12
                        ]
                        data_rows.append(row)

        # Write all rows (note: sheet name has trailing space!)
        self._write_sheet_data(workbook, '6_DL_QuanLy_HanNgach ', data_rows)

    @http.route('/document_extractor/overview_dashboard_data', type='json', auth='user', methods=['POST'])
    def get_overview_dashboard_data(self):
        """
        Get aggregated data for overview dashboard

        Returns:
            dict: Dashboard data with KPIs, charts, and recent activity
        """
        try:
            env = request.env

            # 1. Get data from substance.aggregate (ALREADY EXISTS!)
            SubstanceAggregate = env['substance.aggregate'].sudo()
            aggregate_data = SubstanceAggregate.get_dashboard_data()
            # Returns: {kpis: {...}, charts: {...}, details: {...}}

            # 2. Add document counts
            Document = env['document.extraction'].sudo()
            total_docs = Document.search_count([])
            form01_count = Document.search_count([('document_type', '=', '01')])
            form02_count = Document.search_count([('document_type', '=', '02')])

            # 3. Add status counts (simplified - 3 statuses)
            draft_count = Document.search_count([('state', '=', 'draft')])
            validated_count = Document.search_count([('state', '=', 'validated')])
            completed_count = Document.search_count([('state', '=', 'completed')])

            # 4. Get recent activity from logs
            Log = env['google.drive.extraction.log'].sudo()
            logs = Log.search([], order='create_date desc', limit=5)

            recent_activity = []
            for log in logs:
                # Format relative time
                from odoo.fields import Datetime

                create_date = log.create_date
                if create_date:
                    now = Datetime.now()
                    diff = now - create_date

                    if diff.days > 0:
                        time_str = _("%d days ago") % diff.days
                    elif diff.seconds >= 3600:
                        hours = diff.seconds // 3600
                        time_str = _("%d hours ago") % hours
                    elif diff.seconds >= 60:
                        minutes = diff.seconds // 60
                        time_str = _("%d mins ago") % minutes
                    else:
                        time_str = _("just now")
                else:
                    time_str = _("Unknown")

                # Get document info if extraction record exists
                document_id = False
                display_name = log.file_name

                if log.extraction_record_id:
                    # Has document: use document name
                    document_id = log.extraction_record_id.id
                    display_name = log.extraction_record_id.name or log.file_name

                recent_activity.append({
                    'id': log.id,
                    'log_id': log.id,  # Always available for clicking log record
                    'document_id': document_id,  # Document ID if exists, False otherwise
                    'action': _("Processed %s") % display_name,
                    'user': log.create_uid.name if log.create_uid else _('System'),
                    'time': time_str,
                    'type': 'upload' if log.status == 'success' else 'processing',
                    'display_name': display_name  # For tooltip
                })

            # 5. Calculate Top Substances (optimized SQL query)
            # Use direct SQL query to aggregate and sort in database
            # Extract translated name from JSONB field based on user's language
            user_lang = env.user.lang or 'en_US'
            env.cr.execute("""
                SELECT
                    cs.id,
                    COALESCE(
                        cs.name->>%s,
                        cs.name->>'en_US',
                        cs.name::text
                    ) as translated_name,
                    SUM(COALESCE(sa.total_co2e, 0)) as total_co2e
                FROM substance_aggregate sa
                INNER JOIN controlled_substance cs ON sa.substance_id = cs.id
                WHERE sa.substance_id IS NOT NULL
                GROUP BY cs.id, cs.name
                ORDER BY total_co2e DESC
                LIMIT 5
            """, (user_lang,))

            top_substances_raw = env.cr.fetchall()
            top_substances = [
                {
                    'name': row[1],
                    'value': row[2]
                }
                for row in top_substances_raw
            ]

            # 6. Merge data and return
            return {
                'error': False,
                'kpis': {
                    # From substance.aggregate
                    'total_usage_kg': aggregate_data['kpis']['total_usage_kg'],
                    'total_co2e': aggregate_data['kpis']['total_co2e'],
                    'organization_count': aggregate_data['kpis']['organization_count'],
                    'document_count': aggregate_data['kpis']['document_count'],
                    # Additional counts
                    'total_docs': total_docs,
                    'form01_count': form01_count,
                    'form02_count': form02_count,
                    'status_counts': {
                        'draft': draft_count,
                        'validated': validated_count,
                        'completed': completed_count
                    }
                },
                'charts': {
                    'trend_by_year': aggregate_data['charts']['trend_by_year'],
                    'top_substances': top_substances,  # Changed from top_companies
                    'by_activity_type': aggregate_data['charts']['by_activity_type'],
                },
                'recent_activity': recent_activity
            }

        except Exception as e:
            _logger.error(f'Error fetching overview dashboard data: {str(e)}', exc_info=True)
            return {
                'error': True,
                'message': str(e)
            }
